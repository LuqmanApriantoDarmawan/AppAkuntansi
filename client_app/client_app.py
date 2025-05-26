import sys
import mysql.connector
from fpdf import FPDF
import datetime
import os
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
                            QGridLayout, QLabel, QPushButton, QTreeWidget, QTreeWidgetItem,
                            QScrollArea, QDialog, QRadioButton, QButtonGroup, QLineEdit,
                            QTextEdit, QMessageBox, QFrame, QSpacerItem, QSizePolicy, QInputDialog)
from PyQt5.QtCore import Qt, QSize, pyqtSignal
from PyQt5.QtGui import QFont, QIcon, QColor, QPixmap, QPalette

# Database configuration
DB_CONFIG = {
    'host': 'localhost',
    'user': 'root',
    'password': '',
    'database': 'akuntansi'
}

# Email configuration
EMAIL_CONFIG = {
    'sender': 'your_email@gmail.com',
    'password': 'your_email_password',
    'receiver': 'coddulu977@gmail.com',
    'smtp_server': 'smtp.gmail.com',
    'smtp_port': 587
}

def connect_db():
    return mysql.connector.connect(**DB_CONFIG)

class ProductCard(QPushButton):
    addedToCart = pyqtSignal(str, str, float)  # (code, name, price)

    def __init__(self, product_data, parent=None):
        super().__init__(parent)
        self.product_data = product_data
        self.setup_ui()
        
    def setup_ui(self):
        self.setFixedSize(220, 220)
        self.setCursor(Qt.PointingHandCursor)
        self.setStyleSheet("""
            QPushButton {
                background: #FFFFFF;
                border-radius: 15px;
                border: 2px solid #e0e0e0;
                padding: 15px;
            }
            QPushButton:hover {
                border-color: #3498db;
                background: #f8f9fa;
            }
        """)
        
        layout = QVBoxLayout(self)
        layout.setContentsMargins(10, 15, 10, 15)
        layout.setSpacing(8)
        
        # Product name
        lbl_name = QLabel(self.product_data['nama'])
        lbl_name.setFont(QFont('Arial', 12, QFont.Bold))
        lbl_name.setAlignment(Qt.AlignCenter)
        lbl_name.setWordWrap(True)
        lbl_name.setStyleSheet("color: #2d3436; margin-bottom: 10px;")
        
        # Price
        lbl_price = QLabel(f"Rp {self.product_data['harga_jual']:,}")
        lbl_price.setFont(QFont('Arial', 14, QFont.Bold))
        lbl_price.setStyleSheet("color: #27ae60;")
        lbl_price.setAlignment(Qt.AlignCenter)
        
        # Stock
        lbl_stock = QLabel(f"Stok: {self.product_data['stok']}")
        lbl_stock.setFont(QFont('Arial', 10))
        lbl_stock.setStyleSheet("color: #636e72;")
        lbl_stock.setAlignment(Qt.AlignCenter)
        
        layout.addWidget(lbl_name)
        layout.addWidget(lbl_price)
        layout.addWidget(lbl_stock)
        layout.addStretch()
        
        self.clicked.connect(self.on_click)
        
    def on_click(self):
        self.addedToCart.emit(
            self.product_data['kode'],
            self.product_data['nama'],
            self.product_data['harga_jual']
        )

class ModernPOS(QMainWindow):
    def __init__(self):
        super().__init__()
        self.cart = {}
        self.customer_name = "Self-Service"
        self.current_order_number = self.generate_order_number()
        self.transactions_file = "transactions.xlsx"
        self.initialize_excel_file()
        
        self.setup_ui()
        self.load_products()
        
    def setup_ui(self):
        self.setWindowTitle("Modern POS - Self Service")
        self.setGeometry(100, 100, 1366, 768)
        self.setWindowIcon(QIcon('store_icon.png'))
        
        # Main widget
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        main_layout = QHBoxLayout(main_widget)
        main_layout.setContentsMargins(20, 20, 20, 20)
        main_layout.setSpacing(20)
        
        # Left panel - Products
        left_panel = QFrame()
        left_panel.setStyleSheet("background: white; border-radius: 20px;")
        left_layout = QVBoxLayout(left_panel)
        left_layout.setContentsMargins(15, 15, 15, 15)
        left_layout.setSpacing(15)
        
        # Product grid
        self.product_scroll = QScrollArea()
        self.product_scroll.setWidgetResizable(True)
        self.product_container = QWidget()
        self.product_grid = QGridLayout(self.product_container)
        self.product_container.setLayout(self.product_grid)
        self.product_scroll.setWidget(self.product_container)
        
        left_layout.addWidget(QLabel("Daftar Produk", styleSheet="font-size: 18pt; color: #2d3436; font-weight: bold; padding-bottom: 10px;"))
        left_layout.addWidget(self.product_scroll)
        
        # Right panel - Cart
        right_panel = QFrame()
        right_panel.setStyleSheet("""
            QFrame {
                background: #f8f9fa;
                border-radius: 20px;
                border: 2px solid #e0e0e0;
            }
        """)
        right_layout = QVBoxLayout(right_panel)
        right_layout.setContentsMargins(20, 20, 20, 20)
        right_layout.setSpacing(15)
        
        # Cart header
        cart_header = QLabel("Keranjang Belanja")
        cart_header.setStyleSheet("font-size: 20pt; color: #2d3436; font-weight: bold; padding-bottom: 10px;")
        right_layout.addWidget(cart_header)
        
        # Cart tree
        self.cart_tree = QTreeWidget()
        self.cart_tree.setHeaderLabels(["Produk", "Qty", "Harga Satuan", "Total"])
        self.cart_tree.setStyleSheet("""
            QTreeWidget {
                font-size: 12pt;
                background: white;
                border-radius: 10px;
                padding: 10px;
                border: 1px solid #dcdde1;
            }
            QHeaderView::section {
                background: #3498db;
                color: white;
                padding: 10px;
                border: none;
                font-weight: bold;
            }
        """)
        self.cart_tree.setColumnWidth(0, 280)
        self.cart_tree.setColumnWidth(1, 80)
        self.cart_tree.setColumnWidth(2, 150)
        self.cart_tree.setColumnWidth(3, 150)
        right_layout.addWidget(self.cart_tree)
        
        # Total section
        total_frame = QFrame()
        total_frame.setStyleSheet("background: white; border-radius: 10px; padding: 15px;")
        total_layout = QVBoxLayout(total_frame)
        
        self.lbl_subtotal = QLabel("Subtotal: Rp 0")
        self.lbl_subtotal.setStyleSheet("font-size: 14pt; color: #636e72;")
        self.lbl_subtotal.setAlignment(Qt.AlignRight)
        
        self.lbl_total = QLabel("Total: Rp 0")
        self.lbl_total.setStyleSheet("font-size: 16pt; color: #27ae60; font-weight: bold;")
        self.lbl_total.setAlignment(Qt.AlignRight)
        
        total_layout.addWidget(self.lbl_subtotal)
        total_layout.addWidget(self.lbl_total)
        right_layout.addWidget(total_frame)
        
        # Buttons
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(10)
        
        self.btn_clear = QPushButton("Hapus Keranjang")
        self.btn_clear.setIcon(QIcon('clear.png'))
        self.btn_clear.setStyleSheet(self.get_button_style("#e74c3c", 14))
        self.btn_clear.clicked.connect(self.clear_cart)
        
        self.btn_pay = QPushButton("Proses Pembayaran")
        self.btn_pay.setIcon(QIcon('payment.png'))
        self.btn_pay.setStyleSheet(self.get_button_style("#2ecc71", 16))
        self.btn_pay.clicked.connect(self.process_payment)
        
        btn_layout.addWidget(self.btn_clear)
        btn_layout.addWidget(self.btn_pay)
        right_layout.addLayout(btn_layout)
        
        main_layout.addWidget(left_panel, 65)
        main_layout.addWidget(right_panel, 35)
        
    def get_button_style(self, color, font_size):
        return f"""
            QPushButton {{
                background-color: {color};
                color: white;
                border: none;
                padding: 12px 24px;
                border-radius: 8px;
                font-size: {font_size}pt;
            }}
            QPushButton:hover {{
                background-color: {self.darken_color(color)};
            }}
        """
        
    def darken_color(self, hex_color, factor=0.8):
        color = QColor(hex_color)
        return color.darker(100 + int(100 * (1 - factor)))

    def initialize_excel_file(self):
        if not os.path.exists(self.transactions_file):
            wb = Workbook()
            ws = wb.active
            # ... (same as original Excel initialization)

    def generate_order_number(self):
        now = datetime.datetime.now()
        return f"INV-{now.strftime('%Y%m%d-%H%M%S')}"

    def load_products(self):
        # Clear existing products
        for i in reversed(range(self.product_grid.count())): 
            self.product_grid.itemAt(i).widget().deleteLater()
            
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("SELECT kode, nama, harga_jual, stok FROM barang WHERE stok > 0")
        products = cursor.fetchall()
        conn.close()
        
        row = col = 0
        for kode, nama, harga, stok in products:
            product_data = {
                'kode': kode,
                'nama': nama,
                'harga_jual': harga,
                'stok': stok
            }
            card = ProductCard(product_data)
            card.addedToCart.connect(self.add_to_cart)
            self.product_grid.addWidget(card, row, col)
            col += 1
            if col > 3:
                col = 0
                row += 1

    def add_to_cart(self, product_code, product_name, price):
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("SELECT stok FROM barang WHERE kode = %s", (product_code,))
        current_stock = cursor.fetchone()[0]
        conn.close()
        
        if current_stock <= 0:
            QMessageBox.warning(self, "Stok Habis", f"Stok {product_name} sudah habis!")
            self.load_products()
            return
        
        if product_code in self.cart:
            if self.cart[product_code]['qty'] >= current_stock:
                QMessageBox.warning(self, "Stok Tidak Cukup", 
                                  f"Stok {product_name} tersisa: {current_stock}")
                return
            self.cart[product_code]['qty'] += 1
        else:
            self.cart[product_code] = {
                'name': product_name,
                'price': price,
                'qty': 1
            }
            
        self.update_cart_display()

    def update_cart_display(self):
        self.cart_tree.clear()
        subtotal = 0
        
        for code, item in self.cart.items():
            total = item['price'] * item['qty']
            subtotal += total
            
            tree_item = QTreeWidgetItem(self.cart_tree)
            tree_item.setText(0, item['name'])
            tree_item.setText(1, str(item['qty']))
            tree_item.setText(2, f"Rp {item['price']:,}")
            tree_item.setText(3, f"Rp {total:,}")
            
        self.lbl_subtotal.setText(f"Subtotal: Rp {subtotal:,}")
        self.lbl_total.setText(f"Total: Rp {subtotal:,}")

    def clear_cart(self):
        reply = QMessageBox.question(
            self, 'Konfirmasi',
            'Yakin ingin mengosongkan keranjang belanja?',
            QMessageBox.Yes | QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            self.cart = {}
            self.update_cart_display()

    def process_payment(self):
        if not self.cart:
            QMessageBox.warning(self, "Keranjang Kosong", "Tambahkan produk terlebih dahulu!")
            return
        
        # Get customer name
        name, ok = QInputDialog.getText(
            self, 'Nama Pelanggan',
            'Masukkan nama pelanggan (opsional):',
            QLineEdit.Normal,
            self.customer_name
        )
        if ok:  # Hanya update jika user klik OK
            self.customer_name = name.strip() or "Self-Service"
        else:  # Jika user cancel, jangan lanjutkan proses
            return
            
        try:
            total = self.calculate_total()
            payment_dialog = PaymentDialog(total)
            if payment_dialog.exec_() == QDialog.Accepted:
                self.finalize_payment(payment_dialog.payment_method)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal memproses pembayaran: {str(e)}")
            
    def calculate_total(self):
        return sum(item['price'] * item['qty'] for item in self.cart.values())

    def finalize_payment(self, method):
        try:
            conn = connect_db()
            cursor = conn.cursor()
            
            # Update database
            for code, item in self.cart.items():
                cursor.execute("UPDATE barang SET stok = stok - %s WHERE kode = %s",
                             (item['qty'], code))
                
                cursor.execute("""
                    INSERT INTO penjualan (
                        tanggal, pelanggan, kode_barang, 
                        jumlah, harga_satuan, total
                    ) VALUES (%s, %s, %s, %s, %s, %s)
                """, (
                    datetime.datetime.now(),
                    self.customer_name,
                    code,
                    item['qty'],
                    item['price'],
                    item['price'] * item['qty']
                ))
                
            conn.commit()
            conn.close()
            
            # Generate receipt
            self.generate_receipt(method)
            
            # Record transaction
            self.record_transaction(method)
            
            # Show success
            QMessageBox.information(
                self, "Pembayaran Berhasil",
                f"Pembayaran sebesar Rp {self.calculate_total():,} berhasil!\n"
                f"Struk telah disimpan di folder receipts."
            )
            
            # Show review dialog
            review_dialog = ReviewDialog()
            if review_dialog.exec_() == QDialog.Accepted:
                self.handle_review(
                    review_dialog.rating,
                    review_dialog.comment,
                    review_dialog.email
                )
            
            self.clear_cart()
            
        except mysql.connector.Error as err:
            QMessageBox.critical(self, "Database Error", f"Error database: {err}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error umum: {str(e)}")
        finally:
            if 'conn' in locals() and conn.is_connected():
                conn.close()

    def generate_receipt(self, method):
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", 'B', 16)
        
        # Header
        pdf.cell(0, 10, "TOKO MODERN", ln=1, align='C')
        pdf.set_font("Arial", '', 10)
        pdf.cell(0, 7, datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S'), ln=1, align='C')
        
        # Customer info
        pdf.set_font("Arial", '', 12)
        pdf.cell(0, 10, f"No. Pesanan: {self.current_order_number}", ln=1)
        pdf.cell(0, 10, f"Pelanggan: {self.customer_name}", ln=1)
        
        # Items
        pdf.set_font("Arial", 'B', 10)
        pdf.cell(100, 10, "Produk", border='B')
        pdf.cell(30, 10, "Qty", border='B')
        pdf.cell(30, 10, "Harga", border='B')
        pdf.cell(30, 10, "Total", border='B', ln=1)
        
        pdf.set_font("Arial", '', 10)
        for item in self.cart.values():
            pdf.cell(100, 10, item['name'])
            pdf.cell(30, 10, str(item['qty']), align='R')
            pdf.cell(30, 10, f"Rp {item['price']:,}", align='R')
            pdf.cell(30, 10, f"Rp {item['price']*item['qty']:,}", align='R', ln=1)
        
        # Total
        pdf.set_font("Arial", 'B', 12)
        pdf.cell(160, 10, "TOTAL:", align='R')
        pdf.cell(30, 10, f"Rp {self.calculate_total():,}", align='R', ln=1)
        
        # Save PDF
        receipt_dir = os.path.join(os.getcwd(), "receipts")
        os.makedirs(receipt_dir, exist_ok=True)
        filename = os.path.join(receipt_dir, f"receipt_{self.current_order_number}.pdf")
        pdf.output(filename)
        
    def record_transaction(self, method):
        try:
            wb = load_workbook(self.transactions_file)
            ws = wb.active
            
            transaction_time = datetime.datetime.now()
            payment_method = "QRIS" if method == "qris" else "Kartu"
            
            for code, item in self.cart.items():
                ws.append([
                    self.current_order_number,
                    transaction_time.strftime('%Y-%m-%d %H:%M:%S'),
                    self.customer_name,
                    "",  # Email akan diupdate nanti
                    code,
                    item['name'],
                    item['qty'],
                    item['price'],
                    item['price'] * item['qty'],
                    payment_method,
                    "",  # Rating
                    ""   # Komentar
                ])
            
            wb.save(self.transactions_file)
            
        except Exception as e:
            QMessageBox.warning(self, "Peringatan", 
                              f"Gagal menyimpan ke Excel: {str(e)}")

    def handle_review(self, rating, comment, email):
        try:
            # Update Excel dengan review
            wb = load_workbook(self.transactions_file)
            ws = wb.active
            
            for row in range(2, ws.max_row + 1):
                if ws.cell(row=row, column=1).value == self.current_order_number:
                    ws.cell(row=row, column=4).value = email
                    ws.cell(row=row, column=11).value = rating
                    ws.cell(row=row, column=12).value = comment
                    
                    if rating >= 4:
                        for col in range(1, 13):
                            ws.cell(row=row, column=col).fill = PatternFill(
                                start_color="C6EFCE", 
                                end_color="C6EFCE", 
                                fill_type="solid")
            
            wb.save(self.transactions_file)
            
            # Kirim email jika ada alamat email
            if email:
                self.send_receipt_email(email, rating, comment)
                
        except Exception as e:
            QMessageBox.warning(self, "Peringatan", 
                              f"Gagal menyimpan review: {str(e)}")

    def send_receipt_email(self, email, rating, comment):
        try:
            msg = MIMEMultipart()
            msg['From'] = EMAIL_CONFIG['sender']
            msg['To'] = email
            msg['Subject'] = f"Struk Pembelian - {self.current_order_number}"
            
            body = f"""
            Terima kasih telah berbelanja di Toko Modern!
            
            Detail Pembelian:
            No. Pesanan: {self.current_order_number}
            Tanggal: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')}
            Pelanggan: {self.customer_name}
            
            Item yang dibeli:
            """
            for item in self.cart.values():
                body += f"- {item['name']} x{item['qty']} @ Rp {item['price']:,}\n"
            
            body += f"""
            Total Pembayaran: Rp {self.calculate_total():,}
            
            Ulasan Anda:
            Rating: {rating}/5
            Komentar: {comment or '-'}
            
            Terima kasih atas ulasannya!
            """
            
            msg.attach(MIMEText(body, 'plain'))
            
            with smtplib.SMTP(EMAIL_CONFIG['smtp_server'], EMAIL_CONFIG['smtp_port']) as server:
                server.starttls()
                server.login(EMAIL_CONFIG['sender'], EMAIL_CONFIG['password'])
                server.send_message(msg)
                
        except Exception as e:
            QMessageBox.warning(self, "Peringatan", 
                              f"Email gagal dikirim: {str(e)}")

class PaymentDialog(QDialog):
    def __init__(self, total):
        super().__init__()
        self.setWindowModality(Qt.ApplicationModal)
        self.total = total
        self.setup_ui()
        self.payment_method = "qris"
        
    @property
    def payment_method(self):
        return "qris" if self.rb_qris.isChecked() else "card"
    
    @payment_method.setter  # Tambahkan setter
    def payment_method(self, value):
        self._payment_method = value
        if value == "qris":
            self.rb_qris.setChecked(True)
        else:
            self.rb_card.setChecked(True)
        
    def setup_ui(self):
        self.setWindowTitle("Pembayaran")
        self.setFixedSize(400, 320)
        self.setStyleSheet("""
            QDialog {
                background: white;
                border-radius: 15px;
                padding: 20px;
            }
        """)
        
        layout = QVBoxLayout(self)
        layout.setContentsMargins(30, 20, 30, 20)
        layout.setSpacing(15)
        
        lbl_title = QLabel("Metode Pembayaran")
        lbl_title.setStyleSheet("font-size: 18pt; color: #2d3436; font-weight: bold;")
        lbl_title.setAlignment(Qt.AlignCenter)
        
        self.rb_qris = QRadioButton("QRIS (Gopay, OVO, Dana)")
        self.rb_card = QRadioButton("Kartu Kredit/Debit")
        self.rb_qris.setChecked(True)
        
        style = """
            QRadioButton {
                font-size: 14pt;
                padding: 12px;
                spacing: 10px;
                border-radius: 8px;
                background: #f8f9fa;
            }
            QRadioButton::indicator {
                width: 20px;
                height: 20px;
            }
        """
        self.rb_qris.setStyleSheet(style)
        self.rb_card.setStyleSheet(style)
        
        # Total Display
        lbl_total = QLabel(f"Total: Rp {self.total:,}")
        lbl_total.setStyleSheet("""
            font-size: 16pt; 
            color: #27ae60; 
            font-weight: bold;
            padding: 10px;
            background: #f8f9fa;
            border-radius: 8px;
        """)
        lbl_total.setAlignment(Qt.AlignCenter)
        
        # Confirm Button
        btn_confirm = QPushButton("Konfirmasi Pembayaran")
        btn_confirm.setStyleSheet("""
            QPushButton {
                background: #2ecc71;
                color: white;
                padding: 15px;
                border-radius: 8px;
                font-size: 14pt;
            }
            QPushButton:hover {
                background: #27ae60;
            }
        """)
        btn_confirm.clicked.connect(self.accept)
        
        layout.addWidget(lbl_title)
        layout.addWidget(self.rb_qris)
        layout.addWidget(self.rb_card)
        layout.addSpacerItem(QSpacerItem(20, 40, QSizePolicy.Minimum, QSizePolicy.Expanding))
        layout.addWidget(lbl_total)
        layout.addWidget(btn_confirm)

class ReviewDialog(QDialog):
    def __init__(self):
        super().__init__()
        # Remove these lines since they're not needed
        # self.rating = 5  # This is what caused the error
        # self.comment = ""
        # self.email = ""
        self.setup_ui()
        
    def setup_ui(self):
        self.setWindowTitle("Ulasan Pembelian")
        self.setFixedSize(800, 750)
        self.setStyleSheet("""
            QDialog {
                background: white;
                border-radius: 15px;
                padding: 25px;
            }
        """)
        
        layout = QVBoxLayout(self)
        
        lbl_title = QLabel("Beri Ulasan Anda")
        lbl_title.setStyleSheet("font-size: 18pt; color: #2d3436; font-weight: bold;")
        lbl_title.setAlignment(Qt.AlignCenter)
        
        # Email input
        lbl_email = QLabel("Email (untuk notifikasi):")
        self.txt_email = QLineEdit()
        self.txt_email.setPlaceholderText("opsional")
        
        # Rating
        lbl_rating = QLabel("Rating:")
        self.rating_buttons = QButtonGroup()
        rating_layout = QHBoxLayout()
        for i in range(1, 6):
            rb = QRadioButton(str(i))
            rb.setStyleSheet("font-size: 14pt;")
            self.rating_buttons.addButton(rb, i)
            rating_layout.addWidget(rb)
        self.rating_buttons.button(5).setChecked(True)  # Default to 5 stars
        
        # Comments
        lbl_comment = QLabel("Komentar:")
        self.txt_comment = QTextEdit()
        self.txt_comment.setPlaceholderText("Masukkan komentar tambahan...")
        
        # Submit button
        btn_submit = QPushButton("Kirim Ulasan")
        btn_submit.setStyleSheet("""
            QPushButton {
                background: #3498db;
                color: white;
                padding: 15px;
                border-radius: 8px;
                font-size: 14pt;
            }
            QPushButton:hover {
                background: #2980b9;
            }
        """)
        btn_submit.clicked.connect(self.accept)
        
        layout.addWidget(lbl_title)
        layout.addSpacing(20)
        layout.addWidget(lbl_email)
        layout.addWidget(self.txt_email)
        layout.addWidget(lbl_rating)
        layout.addLayout(rating_layout)
        layout.addWidget(lbl_comment)
        layout.addWidget(self.txt_comment)
        layout.addWidget(btn_submit)
        
    @property
    def rating(self):
        return self.rating_buttons.checkedId()
    
    @property
    def comment(self):
        return self.txt_comment.toPlainText().strip()
    
    @property
    def email(self):
        return self.txt_email.text().strip()


if __name__ == '__main__':
    try:
        app = QApplication(sys.argv)
        window = ModernPOS()
        window.show()
        sys.exit(app.exec_())
    except Exception as e:
        print(f"Critical error: {str(e)}")
        QMessageBox.critical(None, "Fatal Error", f"Aplikasi harus ditutup: {str(e)}")
