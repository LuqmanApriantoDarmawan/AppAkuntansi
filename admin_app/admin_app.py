import sys
import mysql.connector
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QTabWidget, QTableWidget, QTableWidgetItem,
                            QLabel, QLineEdit, QPushButton, QVBoxLayout, QHBoxLayout, QFormLayout, 
                            QDateEdit, QStatusBar, QHeaderView, QMessageBox, QFileDialog)
from PyQt5.QtCore import Qt, QDate
from PyQt5.QtGui import QFont, QColor
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image as xlImage
from openpyxl.utils import get_column_letter
from fpdf import FPDF
import tempfile

# Database configuration
DB_CONFIG = {
    'host': 'localhost',
    'user': 'root',
    'password': '',
    'database': 'akuntansi'
}

def connect_db():
    """Membuat koneksi ke database MySQL"""
    return mysql.connector.connect(**DB_CONFIG)

class AdminApp(QMainWindow):
    def __init__(self):
        super().__init__()
        # Variabel untuk menyimpan data laporan
        self.report_data = {
            'total_penjualan': 0,
            'total_pembelian': 0,
            'laba_kotor': 0,
            'laba_bersih': 0,
            'start_date': QDate.currentDate().addDays(-30).toString("yyyy-MM-dd"),
            'end_date': QDate.currentDate().toString("yyyy-MM-dd"),
            'periode': f"{QDate.currentDate().addDays(-30).toString('dd/MM/yyyy')} - {QDate.currentDate().toString('dd/MM/yyyy')}"
        }
        
        # Setup awal UI
        self.setup_ui()
        self.load_products()
        
        # Set tanggal default untuk filter transaksi
        self.date_from.setDate(QDate.currentDate().addDays(-7))
        self.date_to.setDate(QDate.currentDate())
        
        # Load data awal
        self.filter_transactions()
        self.update_report()
        
    def setup_ui(self):
        """Mengatur tampilan utama aplikasi"""
        self.setWindowTitle("Admin Dashboard - Toko Modern")
        self.setGeometry(100, 100, 1200, 800)
        self.setStyleSheet(self.get_stylesheet())
        
        # Create main tab widget
        self.tabs = QTabWidget()
        self.setCentralWidget(self.tabs)
        
        # Create tabs
        self.create_product_tab()
        self.create_transaction_tab()
        self.create_report_tab()
        
        # Create status bar
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        self.status_bar.showMessage("Ready")
    
    def get_stylesheet(self):
        """Mengembalikan stylesheet untuk aplikasi"""
        return """
            QMainWindow {
                background-color: #f0f0f0;
            }
            QTabWidget::pane {
                border: 1px solid #d0d0d0;
                background: white;
            }
            QTabBar::tab {
                background: #e0e0e0;
                border: 1px solid #d0d0d0;
                padding: 8px 12px;
                font-weight: bold;
                color: #404040;
            }
            QTabBar::tab:selected {
                background: white;
                border-bottom: 2px solid #2196F3;
                color: #2196F3;
            }
            QTableWidget {
                background: white;
                border: 1px solid #d0d0d0;
                font-size: 12px;
            }
            QHeaderView::section {
                background-color: #2196F3;
                color: white;
                font-weight: bold;
                padding: 6px;
                border: none;
            }
            QPushButton {
                padding: 8px 12px;
                border-radius: 4px;
                font-weight: bold;
            }
            QLineEdit {
                padding: 6px;
                border: 1px solid #d0d0d0;
                border-radius: 4px;
            }
        """
        
    def create_product_tab(self):
        """Membuat tab untuk manajemen produk"""
        product_tab = QWidget()
        layout = QVBoxLayout(product_tab)
        
        # Form layout untuk input produk
        form_layout = QFormLayout()
        
        # Input fields
        self.kode_input = QLineEdit()
        self.nama_input = QLineEdit()
        self.stok_input = QLineEdit()
        self.harga_beli_input = QLineEdit()
        self.harga_jual_input = QLineEdit()
        
        form_layout.addRow("Kode:", self.kode_input)
        form_layout.addRow("Nama:", self.nama_input)
        form_layout.addRow("Stok:", self.stok_input)
        form_layout.addRow("Harga Beli:", self.harga_beli_input)
        form_layout.addRow("Harga Jual:", self.harga_jual_input)
        
        # Button layout
        btn_layout = QHBoxLayout()
        self.tambah_btn = QPushButton("Tambah")
        self.update_btn = QPushButton("Update")
        self.hapus_btn = QPushButton("Hapus")
        self.clear_btn = QPushButton("Bersihkan")
        
        # Connect buttons to functions
        self.tambah_btn.clicked.connect(self.tambah_barang)
        self.update_btn.clicked.connect(self.update_barang)
        self.hapus_btn.clicked.connect(self.hapus_barang)
        self.clear_btn.clicked.connect(self.clear_form)
        
        # Add buttons to layout
        btn_layout.addWidget(self.tambah_btn)
        btn_layout.addWidget(self.update_btn)
        btn_layout.addWidget(self.hapus_btn)
        btn_layout.addWidget(self.clear_btn)
        
        # Set button styles
        self.set_button_styles()
        
        # Product table
        self.product_table = QTableWidget()
        self.setup_product_table()
        
        # Add widgets to layout
        layout.addLayout(form_layout)
        layout.addLayout(btn_layout)
        layout.addWidget(self.product_table)
        
        self.tabs.addTab(product_tab, "Manajemen Produk")
    
    def set_button_styles(self):
        """Mengatur style untuk tombol-tombol"""
        self.tambah_btn.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        
        self.update_btn.setStyleSheet("""
            QPushButton {
                background-color: #2196F3;
                color: white;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
        """)
        
        self.hapus_btn.setStyleSheet("""
            QPushButton {
                background-color: #f44336;
                color: white;
            }
            QPushButton:hover {
                background-color: #d32f2f;
            }
        """)
        
        self.clear_btn.setStyleSheet("""
            QPushButton {
                background-color: #9E9E9E;
                color: white;
            }
            QPushButton:hover {
                background-color: #757575;
            }
        """)
    
    def setup_product_table(self):
        """Mengatur tabel produk"""
        self.product_table.setColumnCount(5)
        self.product_table.setHorizontalHeaderLabels(["Kode", "Nama", "Stok", "Harga Beli", "Harga Jual"])
        self.product_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.product_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.product_table.cellClicked.connect(self.on_product_selected)
    
    def create_transaction_tab(self):
        """Membuat tab untuk transaksi"""
        transaction_tab = QWidget()
        layout = QVBoxLayout(transaction_tab)
        
        # Filter layout
        filter_layout = QHBoxLayout()
        self.date_from = QDateEdit()
        self.date_from.setCalendarPopup(True)
        self.date_to = QDateEdit()
        self.date_to.setCalendarPopup(True)
        filter_btn = QPushButton("Filter")
        filter_btn.clicked.connect(self.filter_transactions)
        
        filter_btn.setStyleSheet("""
            QPushButton {
                background-color: #FF9800;
                color: white;
            }
            QPushButton:hover {
                background-color: #F57C00;
            }
        """)
        
        filter_layout.addWidget(QLabel("Dari:"))
        filter_layout.addWidget(self.date_from)
        filter_layout.addWidget(QLabel("Sampai:"))
        filter_layout.addWidget(self.date_to)
        filter_layout.addWidget(filter_btn)
        
        # Transaction table
        self.transaction_table = QTableWidget()
        self.setup_transaction_table()
        
        # Add widgets to layout
        layout.addLayout(filter_layout)
        layout.addWidget(self.transaction_table)
        
        self.tabs.addTab(transaction_tab, "Transaksi")
    
    def setup_transaction_table(self):
        """Mengatur tabel transaksi"""
        self.transaction_table.setColumnCount(8)
        self.transaction_table.setHorizontalHeaderLabels(
            ["ID", "Tanggal", "Kode", "Nama", "Qty", "Harga", "Total", "Keterangan"]
        )
        self.transaction_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
    
    def create_report_tab(self):
        """Membuat tab untuk laporan"""
        report_tab = QWidget()
        layout = QVBoxLayout(report_tab)
        
        # Filter layout untuk laporan
        report_filter_layout = QHBoxLayout()
        self.report_date_from = QDateEdit()
        self.report_date_from.setCalendarPopup(True)
        self.report_date_to = QDateEdit()
        self.report_date_to.setCalendarPopup(True)
        report_filter_btn = QPushButton("Tampilkan Laporan")
        report_filter_btn.clicked.connect(self.update_report)
        
        report_filter_btn.setStyleSheet("""
            QPushButton {
                background-color: #9C27B0;
                color: white;
            }
            QPushButton:hover {
                background-color: #7B1FA2;
            }
        """)
        
        # Set tanggal default untuk laporan
        self.report_date_from.setDate(QDate.currentDate().addDays(-30))
        self.report_date_to.setDate(QDate.currentDate())

        # Update report_data dengan tanggal default
        self.report_data.update({
            'start_date': self.report_date_from.date().toString("yyyy-MM-dd"),
            'end_date': self.report_date_to.date().toString("yyyy-MM-dd"),
            'periode': f"{self.report_date_from.date().toString('dd/MM/yyyy')} - {self.report_date_to.date().toString('dd/MM/yyyy')}"
        })
        
        # Layout untuk tombol ekspor
        export_layout = QHBoxLayout()
        self.setup_export_buttons(export_layout)
        
        # Add filter components to layout
        report_filter_layout.addWidget(QLabel("Periode Laporan:"))
        report_filter_layout.addWidget(self.report_date_from)
        report_filter_layout.addWidget(QLabel("s/d"))
        report_filter_layout.addWidget(self.report_date_to)
        report_filter_layout.addWidget(report_filter_btn)
        report_filter_layout.addStretch()
        
        # Graph canvas
        self.figure = Figure(figsize=(8, 6))
        self.canvas = FigureCanvas(self.figure)
        
        # Profit/Loss table
        self.profit_table = QTableWidget()
        self.setup_profit_table()
        
        # Add widgets to layout
        layout.addLayout(report_filter_layout)
        layout.addWidget(self.canvas)
        layout.addWidget(self.profit_table)
        layout.addLayout(export_layout)
        
        self.tabs.addTab(report_tab, "Laporan")
    
    def setup_export_buttons(self, layout):
        """Mengatur tombol ekspor"""
        self.export_excel_btn = QPushButton("📈 Export Excel")
        self.export_pdf_btn = QPushButton("📄 Export PDF")
        
        self.export_excel_btn.clicked.connect(self.export_to_excel)
        self.export_pdf_btn.clicked.connect(self.export_to_pdf)
        
        self.export_excel_btn.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        
        self.export_pdf_btn.setStyleSheet("""
            QPushButton {
                background-color: #2196F3;
                color: white;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
        """)
        
        layout.addWidget(self.export_excel_btn)
        layout.addWidget(self.export_pdf_btn)
        layout.addStretch()
    
    def setup_profit_table(self):
        """Mengatur tabel laba rugi"""
        self.profit_table.setColumnCount(4)
        self.profit_table.setHorizontalHeaderLabels(
            ["Total Penjualan", "Total Pembelian", "Laba Kotor", "Laba Bersih"]
        )
        self.profit_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        
        self.profit_table.setStyleSheet("""
            QTableWidget {
                font-size: 14px;
                font-weight: bold;
            }
            QTableWidget::item {
                padding: 10px;
            }
        """)
    
    # ==================== DATABASE OPERATIONS ====================
    
    def load_products(self):
        """Memuat data produk dari database ke tabel"""
        try:
            conn = connect_db()
            cursor = conn.cursor()
            cursor.execute("SELECT kode, nama, stok, harga_beli, harga_jual FROM barang")
            products = cursor.fetchall()
            
            self.product_table.setRowCount(len(products))
            for row_idx, row_data in enumerate(products):
                for col_idx, col_data in enumerate(row_data):
                    item = QTableWidgetItem(str(col_data))
                    item.setFlags(item.flags() ^ Qt.ItemIsEditable)
                    self.product_table.setItem(row_idx, col_idx, item)
                    
            # Set alternating row colors
            self.set_table_row_colors(self.product_table)
                    
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal memuat produk: {str(e)}")
        finally:
            if 'conn' in locals():
                conn.close()
    
    def set_table_row_colors(self, table):
        """Mengatur warna baris alternatif untuk tabel"""
        for row in range(table.rowCount()):
            color = QColor(240, 240, 240) if row % 2 == 0 else QColor(255, 255, 255)
            for col in range(table.columnCount()):
                table.item(row, col).setBackground(color)
    
    def filter_transactions(self):
        """Memfilter transaksi berdasarkan tanggal"""
        try:
            date_from = self.date_from.date().toString("yyyy-MM-dd")
            date_to = self.date_to.date().toString("yyyy-MM-dd")
            
            conn = connect_db()
            cursor = conn.cursor()
            cursor.execute("""
                SELECT p.id, DATE_FORMAT(p.tanggal, '%Y-%m-%d'), p.kode_barang, 
                       COALESCE(b.nama, 'Barang dihapus'), p.jumlah, p.harga_satuan, 
                       (p.jumlah * p.harga_satuan), '-'
                FROM penjualan p
                LEFT JOIN barang b ON p.kode_barang = b.kode
                WHERE p.tanggal BETWEEN %s AND %s
                ORDER BY p.tanggal DESC
            """, (date_from + " 00:00:00", date_to + " 23:59:59"))
            
            transactions = cursor.fetchall()
            
            self.transaction_table.setRowCount(len(transactions))
            for row_idx, row_data in enumerate(transactions):
                for col_idx, col_data in enumerate(row_data):
                    item = QTableWidgetItem(str(col_data))
                    item.setFlags(item.flags() ^ Qt.ItemIsEditable)
                    self.transaction_table.setItem(row_idx, col_idx, item)
                    
            self.update_report()
            self.status_bar.showMessage(f"Menampilkan transaksi dari {date_from} sampai {date_to}")
            
            # Set alternating row colors
            self.set_table_row_colors(self.transaction_table)
                    
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal memfilter transaksi: {str(e)}")
        finally:
            if 'conn' in locals():
                conn.close()
    
    # ==================== REPORT FUNCTIONS ====================
    
    def update_report(self):
        """Memperbarui laporan keuangan berdasarkan periode yang dipilih"""
        try:
            # Clear previous figure
            self.figure.clear()
    
            # Get dates from report tab
            start_date = self.report_date_from.date().toString("yyyy-MM-dd")
            end_date = self.report_date_to.date().toString("yyyy-MM-dd")
    
            # Update report data
            self.report_data.update({
                'start_date': start_date,
                'end_date': end_date,
                'periode': f"{self.report_date_from.date().toString('dd/MM/yyyy')} - {self.report_date_to.date().toString('dd/MM/yyyy')}"
            })
    
            # Get financial data from database
            conn = connect_db()
            cursor = conn.cursor()
        
            # Get total sales
            cursor.execute("""
                SELECT SUM(jumlah * harga_satuan) 
                FROM penjualan 
                WHERE tanggal BETWEEN %s AND %s
            """, (start_date + " 00:00:00", end_date + " 23:59:59"))
            total_penjualan = cursor.fetchone()[0] or 0
        
            # Get total purchase cost
            cursor.execute("""
                SELECT SUM(p.jumlah * b.harga_beli)
                FROM penjualan p
                JOIN barang b ON p.kode_barang = b.kode
                WHERE p.tanggal BETWEEN %s AND %s
            """, (start_date + " 00:00:00", end_date + " 23:59:59"))
            total_pembelian = cursor.fetchone()[0] or 0
        
            # Calculate profits
            laba_kotor = total_penjualan - total_pembelian
            laba_bersih = laba_kotor  # Simplified calculation
        
            # Update report data
            self.report_data.update({
                'total_penjualan': total_penjualan,
                'total_pembelian': total_pembelian,
                'laba_kotor': laba_kotor,
                'laba_bersih': laba_bersih
            })
        
            # Update profit/loss table
            self.update_profit_table(total_penjualan, total_pembelian, laba_kotor, laba_bersih)
                
            # Create bar chart
            self.create_profit_chart(total_penjualan, total_pembelian, laba_kotor)
            
            # Update status bar
            self.status_bar.showMessage(f"Laporan diperbarui untuk periode {self.report_data['periode']}")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal membuat laporan: {str(e)}")
        finally:
            if 'conn' in locals():
                conn.close()
    
    def update_profit_table(self, penjualan, pembelian, laba_kotor, laba_bersih):
        """Memperbarui tabel laba rugi"""
        self.profit_table.setRowCount(1)
        values = [penjualan, pembelian, laba_kotor, laba_bersih]
        for col_idx, value in enumerate(values):
            item = QTableWidgetItem(f"Rp{value:,}")
            item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.profit_table.setItem(0, col_idx, item)
    
    def create_profit_chart(self, penjualan, pembelian, laba_kotor):
        """Membuat grafik laba rugi"""
        # Clear previous figure
        self.figure.clear()
        
        ax = self.figure.add_subplot(111)
        labels = ['Penjualan', 'Pembelian', 'Laba Kotor']
        values = [penjualan, pembelian, laba_kotor]
        
        # Set different colors for each bar
        colors = ['#4CAF50', '#F44336', '#2196F3']
        bars = ax.bar(labels, values, color=colors)
        ax.set_title(f'Laporan Laba Rugi\nPeriode: {self.report_data["periode"]}')
        ax.set_ylabel('Jumlah (Rp)')
        
        # Format y-axis to show currency
        ax.yaxis.set_major_formatter('Rp{x:,.0f}')
        
        # Add value labels on top of each bar
        for bar in bars:
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width()/2., height,
                    f'Rp{height:,.0f}',
                    ha='center', va='bottom', fontsize=10)
        
        # Adjust layout
        self.figure.tight_layout()
        
        # Redraw canvas
        self.canvas.draw()
    
    # ==================== EXPORT FUNCTIONS ====================
    
    def export_to_excel(self):
        """Mengekspor laporan ke file Excel"""
        try:
            # Validate report data
            if not self.report_data['start_date']:
                raise ValueError("Data laporan belum tersedia, silakan tampilkan laporan terlebih dahulu")
            
            # Format periode
            start_date = QDate.fromString(self.report_data['start_date'], "yyyy-MM-dd").toString("dd/MM/yyyy")
            end_date = QDate.fromString(self.report_data['end_date'], "yyyy-MM-dd").toString("dd/MM/yyyy")
            periode = f"{start_date} - {end_date}"
            
            # Get save file path
            file_path, _ = QFileDialog.getSaveFileName(
                self, "Simpan Excel", "", "Excel Files (*.xlsx)"
            )
            
            if not file_path:
                return
                
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Laporan"
            
            # Header
            ws['A1'] = "Laporan Keuangan"
            ws['A1'].font = Font(bold=True, size=14)
            ws.merge_cells('A1:D1')
            
            ws['A2'] = f"Periode: {periode}"
            ws.merge_cells('A2:D2')
            
            # Data
            headers = ["Keterangan", "Jumlah (Rp)"]
            data = [
                ("Total Penjualan", self.report_data['total_penjualan']),
                ("Total Pembelian", self.report_data['total_pembelian']),
                ("Laba Kotor", self.report_data['laba_kotor']),
                ("Laba Bersih", self.report_data['laba_bersih'])
            ]
            
            # Write headers
            for col, header in enumerate(headers, 1):
                ws.cell(row=4, column=col, value=header).font = Font(bold=True)
                
            # Write data
            for row, (label, value) in enumerate(data, 5):
                ws.cell(row=row, column=1, value=label)
                ws.cell(row=row, column=2, value=value).number_format = '#,##0'
                
            # Format columns
            self.adjust_excel_columns(ws)
                
            # Save file
            wb.save(file_path)
            QMessageBox.information(self, "Sukses", "Laporan berhasil diekspor ke Excel!")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal ekspor ke Excel: {str(e)}")
    
    def adjust_excel_columns(self, worksheet):
        """Menyesuaikan lebar kolom Excel"""
        for col_idx in range(1, worksheet.max_column + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            for cell in worksheet[column_letter]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def export_to_pdf(self):
        """Mengekspor laporan ke file PDF"""
        try:
            # Validate report data
            if not self.report_data['start_date']:
                raise ValueError("Data laporan belum tersedia, silakan tampilkan laporan terlebih dahulu")
            
            # Format periode
            start_date = QDate.fromString(self.report_data['start_date'], "yyyy-MM-dd").toString("dd/MM/yyyy")
            end_date = QDate.fromString(self.report_data['end_date'], "yyyy-MM-dd").toString("dd/MM/yyyy")
            periode = f"{start_date} - {end_date}"
             
            # Get save file path
            file_path, _ = QFileDialog.getSaveFileName(
                self, "Simpan PDF", "", "PDF Files (*.pdf)"
            )
            
            if not file_path:
                return
            
            # Save chart as temporary image
            chart_path = self.save_chart_image()
                
            # Create PDF
            pdf = FPDF()
            pdf.add_page()
            pdf.set_auto_page_break(auto=True, margin=15)
            
            # Header
            pdf.set_font("Arial", 'B', 16)
            pdf.cell(0, 10, "Laporan Keuangan", 0, 1, 'C')
            pdf.set_font("Arial", '', 12)
            pdf.cell(0, 10, f"Periode: {periode}", 0, 1, 'C')
            pdf.ln(10)
            
            # Chart image
            pdf.image(chart_path, x=10, w=190)
            pdf.ln(10)
            
            # Data table
            pdf.set_font("Arial", 'B', 12)
            pdf.cell(95, 10, "Keterangan", 1, 0, 'C')
            pdf.cell(95, 10, "Jumlah (Rp)", 1, 1, 'C')
            
            pdf.set_font("Arial", '', 11)
            data = [
                ("Total Penjualan", self.report_data['total_penjualan']),
                ("Total Pembelian", self.report_data['total_pembelian']),
                ("Laba Kotor", self.report_data['laba_kotor']),
                ("Laba Bersih", self.report_data['laba_bersih'])
            ]
            
            for label, value in data:
                pdf.cell(95, 10, label, 1, 0, 'L')
                pdf.cell(95, 10, f"{value:,}", 1, 1, 'R')
                
            # Footer
            pdf.ln(10)
            pdf.set_font("Arial", 'I', 8)
            pdf.cell(0, 10, f"Dicetak pada: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", 0, 0, 'R')
            
            # Save PDF
            pdf.output(file_path)
            QMessageBox.information(self, "Sukses", "Laporan berhasil diekspor ke PDF!")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal ekspor ke PDF: {str(e)}")
        finally:
            # Clean up temporary file
            if 'chart_path' in locals():
                import os
                os.unlink(chart_path)
    
    def save_chart_image(self):
        """Menyimpan grafik sebagai gambar sementara"""
        with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmpfile:
            self.figure.savefig(tmpfile.name, dpi=300, bbox_inches='tight')
            return tmpfile.name
    
    # ==================== PRODUCT MANAGEMENT ====================
    
    def on_product_selected(self, row):
        """Menangani ketika produk dipilih di tabel"""
        self.kode_input.setText(self.product_table.item(row, 0).text())
        self.nama_input.setText(self.product_table.item(row, 1).text())
        self.stok_input.setText(self.product_table.item(row, 2).text())
        self.harga_beli_input.setText(self.product_table.item(row, 3).text())
        self.harga_jual_input.setText(self.product_table.item(row, 4).text())
        
    def clear_form(self):
        """Mengosongkan form input produk"""
        self.kode_input.clear()
        self.nama_input.clear()
        self.stok_input.clear()
        self.harga_beli_input.clear()
        self.harga_jual_input.clear()
        
    def validate_product_input(self):
        """Validasi input produk"""
        fields = {
            "Kode": self.kode_input.text(),
            "Nama": self.nama_input.text(),
            "Stok": self.stok_input.text(),
            "Harga Beli": self.harga_beli_input.text(),
            "Harga Jual": self.harga_jual_input.text()
        }
        
        # Check empty fields
        for field, value in fields.items():
            if not value:
                raise ValueError(f"Field {field} harus diisi!")
            
        # Check numeric values
        try:
            stok = int(fields['Stok'])
            harga_beli = int(fields['Harga Beli'])
            harga_jual = int(fields['Harga Jual'])
        except ValueError:
            raise ValueError("Stok dan harga harus berupa angka!")
        
        # Check price validity
        if harga_jual <= harga_beli:
            raise ValueError("Harga jual harus lebih besar dari harga beli!")
        
        return fields
        
    def tambah_barang(self):
        """Menambahkan produk baru"""
        try:
            fields = self.validate_product_input()
            
            conn = connect_db()
            cursor = conn.cursor()
            
            # Check if product code exists
            cursor.execute("SELECT COUNT(*) FROM barang WHERE kode = %s", (fields['Kode'],))
            if cursor.fetchone()[0] > 0:
                raise ValueError("Kode barang sudah ada!")
            
            # Insert new product
            cursor.execute("""
                INSERT INTO barang (kode, nama, stok, harga_beli, harga_jual)
                VALUES (%s, %s, %s, %s, %s)
            """, (fields['Kode'], fields['Nama'], fields['Stok'], 
                 fields['Harga Beli'], fields['Harga Jual']))
            conn.commit()
            
            self.load_products()
            self.clear_form()
            QMessageBox.information(self, "Sukses", "Barang berhasil ditambahkan!")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))
            
    def update_barang(self):
        """Memperbarui produk yang ada"""
        try:
            selected_row = self.product_table.currentRow()
            if selected_row == -1:
                raise ValueError("Pilih barang yang akan diupdate!")
                
            fields = self.validate_product_input()
            current_kode = self.product_table.item(selected_row, 0).text()
            
            conn = connect_db()
            cursor = conn.cursor()
            
            # Update product
            cursor.execute("""
                UPDATE barang 
                SET kode = %s, nama = %s, stok = %s, 
                    harga_beli = %s, harga_jual = %s 
                WHERE kode = %s
            """, (fields['Kode'], fields['Nama'], fields['Stok'],
                 fields['Harga Beli'], fields['Harga Jual'], current_kode))
            conn.commit()
            
            self.load_products()
            self.clear_form()
            QMessageBox.information(self, "Sukses", "Barang berhasil diupdate!")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))
            
    def hapus_barang(self):
        """Menghapus produk"""
        try:
            selected_row = self.product_table.currentRow()
            if selected_row == -1:
                raise ValueError("Pilih barang yang akan dihapus!")
                
            kode = self.product_table.item(selected_row, 0).text()
            nama = self.product_table.item(selected_row, 1).text()
            
            # Confirmation dialog
            confirm = QMessageBox.question(
                self, "Konfirmasi", 
                f"Apakah Anda yakin ingin menghapus {nama} ({kode})?",
                QMessageBox.Yes | QMessageBox.No
            )
            
            if confirm == QMessageBox.Yes:
                conn = connect_db()
                cursor = conn.cursor()
                cursor.execute("DELETE FROM barang WHERE kode = %s", (kode,))
                conn.commit()
                
                self.load_products()
                self.clear_form()
                QMessageBox.information(self, "Sukses", "Barang berhasil dihapus!")
                
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))

if __name__ == '__main__':
    app = QApplication(sys.argv)
    # Set font untuk seluruh aplikasi
    font = QFont()
    font.setFamily("Segoe UI")
    font.setPointSize(10)
    app.setFont(font)
    window = AdminApp()
    window.show()
    sys.exit(app.exec_())